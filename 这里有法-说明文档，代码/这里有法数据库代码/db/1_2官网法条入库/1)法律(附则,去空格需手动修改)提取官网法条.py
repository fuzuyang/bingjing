from docx import Document
from openpyxl import Workbook
import re

#需修改word_path、excel_path
#输出的excel的地址
excel_path = r"C:\Users\29944\Desktop\137反电信网络诈骗法法条正文.xlsx"  #excel地址
#------------------------------------------------------------------------------
# 读取Word文档
word_path = r"C:\Users\29944\Desktop\中华人民共和国反电信网络诈骗法_20220902.docx"  #word地址

doc = Document(word_path)
full_text = ""
for para in doc.paragraphs:
    full_text += para.text + "\n"
    
# 在处理换行符之前，先将全角空格转换为普通空格
full_text = full_text.replace('\u3000', ' ')  # 将全角空格转换为普通空格

# 关键修改：移除每行开头的空格
lines = full_text.split('\n')
cleaned_lines = []
for line in lines:
    # 移除行首的空格（普通空格和制表符）
    cleaned_line = line.lstrip(' \t')
    cleaned_lines.append(cleaned_line)
full_text = '\n'.join(cleaned_lines)

full_text = re.sub(r"[\r\n]+", "\n", full_text)

# 简化正则表达式模式
chinese_num = r"零一二三四五六七八九十百千万拾〇"
any_blank = r"\s*"

# 结构匹配模式 - 改进法条内容提取
patterns = [
    ("编", rf"^第([{chinese_num}]+)编{any_blank}([^\n]+)", re.MULTILINE),
    ("分编", rf"^第([{chinese_num}]+)分编{any_blank}([^\n]+)", re.MULTILINE), 
    ("章", rf"^第([{chinese_num}]+)章{any_blank}([^\n]+)", re.MULTILINE),
    ("节", rf"^第([{chinese_num}]+)节{any_blank}([^\n]+)", re.MULTILINE),
    ("条", rf"^第([{chinese_num}]+)条(?:[之{chinese_num}]+)?{any_blank}([^\n]*)", re.MULTILINE)  # 修改这里
]

structure_elements = []

# 使用一次性扫描所有结构元素
for elem_type, pattern, flags in patterns:
    matches = list(re.finditer(pattern, full_text, flags))
    for match in matches:
        if elem_type == "条":
            # 对于法条，需要提取完整内容
            start_pos = match.start()
            
            # 找到完整的法条编号（包含"之一"、"之二"等）
            full_article_pattern = rf"^第([{chinese_num}]+)条(?:[之{chinese_num}]+)?"
            full_article_match = re.match(full_article_pattern, full_text[start_pos:])
            
            if full_article_match:
                article_number = full_text[start_pos:start_pos + full_article_match.end()].strip()
            else:
                article_number = f"第{match.group(1)}条"
            
            # 找到下一个标题行开始位置
            next_pattern = rf"^第[{chinese_num}]+(?:条(?:[之{chinese_num}]+)?|编|分编|章|节){any_blank}"
            next_match = re.search(next_pattern, full_text[start_pos + 1:], re.MULTILINE)
            
            if next_match:
                content_end = start_pos + 1 + next_match.start()
            else:
                content_end = len(full_text)
            
            # 提取完整内容
            full_content = full_text[start_pos:content_end].strip()
            
            # 分离标题和内容
            title_match = re.match(rf"^第([{chinese_num}]+)条(?:[之{chinese_num}]+)?{any_blank}", full_content)
            if title_match:
                title_end = title_match.end()
                content = full_content[title_end:].strip()
            else:
                content = full_content
            
            structure_elements.append(("条", article_number, content, start_pos, content_end))
        else:
            # 对于编章节，直接记录
            full_text_rep = f"第{match.group(1)}{elem_type} {match.group(2)}"
            structure_elements.append((elem_type, full_text_rep, match.start(), match.end()))

# 按位置排序
structure_elements.sort(key=lambda x: x[3] if len(x) > 3 else x[2])

print(f"找到 {len([x for x in structure_elements if x[0] == '编'])} 个编")
print(f"找到 {len([x for x in structure_elements if x[0] == '分编'])} 个分编") 
print(f"找到 {len([x for x in structure_elements if x[0] == '章'])} 个章")
print(f"找到 {len([x for x in structure_elements if x[0] == '节'])} 个节")
print(f"找到 {len([x for x in structure_elements if x[0] == '条'])} 个条")

# 构建章节路径
section_data = []
current_bian = ""
current_fenbian = "" 
current_zhang = ""
current_jie = ""

for element in structure_elements:
    elem_type = element[0]
    
    if elem_type == "编":
        current_bian = element[1]
        current_fenbian = ""
        current_zhang = ""
        current_jie = ""
    elif elem_type == "分编":
        current_fenbian = element[1]
        current_zhang = ""
        current_jie = ""
    elif elem_type == "章":
        current_zhang = element[1]
        current_jie = ""
    elif elem_type == "节":
        current_jie = element[1]
    elif elem_type == "条":
        # 构建完整路径
        path_parts = []
        if current_bian:
            path_parts.append(current_bian)
        if current_fenbian:
            path_parts.append(current_fenbian)
        if current_zhang:
            path_parts.append(current_zhang)
        if current_jie:
            path_parts.append(current_jie)
        
        chapter_path = " > ".join(path_parts)
        section_data.append((element[1], element[2], chapter_path))

print(f"\n共提取到 {len(section_data)} 条条文")

# 检查内容是否完整
empty_count = sum(1 for _, content, _ in section_data if not content.strip())
if empty_count > 0:
    print(f"⚠️ 发现 {empty_count} 条法条内容为空")

# 检查是否有重复
article_numbers = [item[0] for item in section_data]
if len(article_numbers) != len(set(article_numbers)):
    duplicates = [item for item in article_numbers if article_numbers.count(item) > 1]
    print(f"❌ 发现重复法条: {set(duplicates)}")
else:
    print("✅ 没有发现重复法条")



# 保存到Excel
wb = Workbook()
ws = wb.active
ws.title = "法条正文"

# 添加表头
headers = ["Chapter", "Provision_number", "Provision_text"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 写入数据
for idx, (num_text, content, chapter_path) in enumerate(section_data, start=2):
    ws.cell(row=idx, column=1, value=chapter_path)
    ws.cell(row=idx, column=2, value=num_text)
    ws.cell(row=idx, column=3, value=content)


wb.save(excel_path)
print(f"\n提取完成！结果已保存到：{excel_path}")
print(f"共保存 {len(section_data)} 条法条到Excel")