from docx import Document
from openpyxl import Workbook
import re

# 需修改word_path、excel_path
# 输出的excel的地址
excel_path=r"C:\Users\29944\Desktop\135枪支、弹药、爆炸物司法解释法条正文.xlsx"  # excel地址
# ------------------------------------------------------------------------------
# 读取Word文档
word_path=r"C:\Users\29944\Desktop\最高人民法院关于审理非法制造、买卖、运输枪支、弹药、爆炸物等刑事案件具体应用法律若干问题的解释_20091116.docx"  # word地址

doc = Document(word_path)
full_text = ""
for para in doc.paragraphs:
    full_text += para.text + "\n"
    
# 在处理换行符之前，先将全角空格转换为普通空格
full_text = full_text.replace('\u3000', ' ')  # 将全角空格转换为普通空格

# 关键修改：移除每行开头的空格
lines = full_text.split('\n')
cleaned_lines = []
for line in lines:
    # 移除行首的空格（普通空格和制表符）
    cleaned_line = line.lstrip(' \t')
    cleaned_lines.append(cleaned_line)
full_text = '\n'.join(cleaned_lines)

full_text = re.sub(r"[\r\n]+", "\n", full_text)

# 简化正则表达式模式
chinese_num = r"零一二三四五六七八九十百千万拾〇"
any_blank = r"\s*"

# 结构匹配模式 - 修改章节结构提取，保留原有的法条提取
patterns = [
    ("条", rf"^第([{chinese_num}]+)条(?:[之{chinese_num}]+)?{any_blank}([^\n]*)", re.MULTILINE)  # 法条模式保持不变
]

structure_elements = []

# 首先提取所有法条
for elem_type, pattern, flags in patterns:
    matches = list(re.finditer(pattern, full_text, flags))
    for match in matches:
        start_pos = match.start()
        
        # 对于法条，需要提取完整内容
        full_article_pattern = rf"^第([{chinese_num}]+)条(?:[之{chinese_num}]+)?"
        full_article_match = re.match(full_article_pattern, full_text[start_pos:])
        
        if full_article_match:
            article_number = full_text[start_pos:start_pos + full_article_match.end()].strip()
        else:
            article_number = f"第{match.group(1)}条"
        
        # 找到下一个标题行开始位置
        # 注意：现在下一个标题可能是"一、xxx"或者"第X条"
        next_pattern = rf"^((?:[一二三四五六七八九十]+、)|(?:第[{chinese_num}]+条(?:[之{chinese_num}]+)?))\s*"
        next_match = re.search(next_pattern, full_text[start_pos + 1:], re.MULTILINE)
        
        if next_match:
            content_end = start_pos + 1 + next_match.start()
        else:
            content_end = len(full_text)
        
        # 提取完整内容
        full_content = full_text[start_pos:content_end].strip()
        
        # 分离标题和内容
        title_match = re.match(rf"^第([{chinese_num}]+)条(?:[之{chinese_num}]+)?{any_blank}", full_content)
        if title_match:
            title_end = title_match.end()
            content = full_content[title_end:].strip()
        else:
            content = full_content
        
        structure_elements.append(("条", article_number, content, start_pos, content_end))

# 按位置排序
structure_elements.sort(key=lambda x: x[3] if len(x) > 3 else x[2])

print(f"找到 {len([x for x in structure_elements if x[0] == '条'])} 个条")

# 构建章节路径 - 这里需要修改逻辑来处理新的章节结构
section_data = []
current_chapter_path = []

# 为每个法条找到其所属的章节
for i, element in enumerate(structure_elements):
    if element[0] == "条":
        article_start = element[3]  # 法条开始位置
        
        # 找到这个法条前面的章节标题
        # 查找在法条之前的所有章节标题（格式：一、xxx 二、xxx 等）
        chapter_pattern = rf"^([一二三四五六七八九十]+)、([^\n]+)"
        chapters_before = []
        
        # 从文件开始到这个法条之前的所有内容中查找章节标题
        text_before_article = full_text[:article_start]
        chapter_matches = re.finditer(chapter_pattern, text_before_article, re.MULTILINE)
        
        for match in chapter_matches:
            chapter_num = match.group(1)
            chapter_title = match.group(2).strip()
            chapter_pos = match.start()
            chapters_before.append((chapter_num, chapter_title, chapter_pos))
        
        # 如果有找到章节标题，使用最接近的一个
        if chapters_before:
            # 按位置排序，找到最接近法条的那个章节
            chapters_before.sort(key=lambda x: x[2], reverse=True)
            latest_chapter = chapters_before[0]
            chapter_path = f"{latest_chapter[0]}、{latest_chapter[1]}"
        else:
            chapter_path = ""
        
        section_data.append((element[1], element[2], chapter_path))

print(f"\n共提取到 {len(section_data)} 条条文")

# 检查内容是否完整
empty_count = sum(1 for _, content, _ in section_data if not content.strip())
if empty_count > 0:
    print(f"⚠️ 发现 {empty_count} 条法条内容为空")

# 检查是否有重复
article_numbers = [item[0] for item in section_data]
if len(article_numbers) != len(set(article_numbers)):
    duplicates = [item for item in article_numbers if article_numbers.count(item) > 1]
    print(f"❌ 发现重复法条: {set(duplicates)}")
else:
    print("✅ 没有发现重复法条")

# 保存到Excel
wb = Workbook()
ws = wb.active
ws.title = "法条正文"

# 添加表头
headers = ["Chapter","Provision_number", "Provision_text"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 写入数据
for idx, (num_text, content, chapter_path) in enumerate(section_data, start=2):
    ws.cell(row=idx, column=1, value=chapter_path)
    ws.cell(row=idx, column=2, value=num_text)
    ws.cell(row=idx, column=3, value=content)

wb.save(excel_path)
print(f"\n提取完成！结果已保存到：{excel_path}")
print(f"共保存 {len(section_data)} 条法条到Excel")